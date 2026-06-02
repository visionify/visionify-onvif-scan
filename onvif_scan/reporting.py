"""Build Excel + self-contained HTML reports and email them via the relay."""
from __future__ import annotations

import base64
from pathlib import Path
from typing import List

from .models import ScanResult


# --------------------------------------------------------------------------- Excel
def build_excel(result: ScanResult, wd: Path, out: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    green = PatternFill("solid", fgColor="C6EFCE")
    amber = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    bold = Font(bold=True)

    def fill_for(status: str):
        return {"good": green, "usable": amber, "failed": red}.get(status, None)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    headers = ["IP", "Manufacturer", "Model", "Firmware", "Streams",
               "Best Res", "Best FPS", "Status", "Notes"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
    for d in result.devices:
        ok = [s for s in d.streams if s.error is None]
        best_res = max((s for s in ok), key=lambda s: s.width * s.height, default=None)
        best_fps = max((s.fps_measured for s in ok), default=0)
        ws.append([
            d.ip, d.manufacturer, d.model, d.firmware, len(d.streams),
            best_res.resolution if best_res else "",
            best_fps, d.status, d.error or "",
        ])
        f = fill_for(d.status)
        if f:
            ws.cell(row=ws.max_row, column=8).fill = f

    # Streams sheet (with embedded thumbnails)
    ss = wb.create_sheet("Streams")
    sheaders = ["Device IP", "Profile", "Resolution", "Codec", "FPS adv",
                "FPS measured", "Bitrate kbps", "Quality", "Exposure",
                "RTSP URL", "Thumbnail"]
    ss.append(sheaders)
    for c in ss[1]:
        c.font = bold
    row = 2
    for d in result.devices:
        for s in d.streams:
            ss.append([
                d.ip, s.profile, s.resolution, s.codec, s.fps_advertised,
                s.fps_measured, s.bitrate_kbps, s.quality_score, s.exposure,
                s.rtsp_url, "",
            ])
            if s.screenshots:
                img_path = wd / s.screenshots[0]
                if img_path.exists():
                    try:
                        xim = XLImage(str(img_path))
                        xim.width, xim.height = 160, 90
                        ss.row_dimensions[row].height = 70
                        ss.add_image(xim, f"K{row}")
                    except Exception:
                        pass
            row += 1

    # Failures sheet
    fs = wb.create_sheet("Failures")
    fs.append(["IP", "Scope", "Reason"])
    for c in fs[1]:
        c.font = bold
    for d in result.devices:
        if d.error or not d.reachable:
            fs.append([d.ip, "device", d.error or "unreachable"])
        for s in d.streams:
            if s.error:
                fs.append([d.ip, f"stream:{s.profile}", s.error])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(50, width + 2)

    wb.save(out)
    return out


# --------------------------------------------------------------------------- HTML
def _img_data_uri(path: Path, max_width: int = 480) -> str:
    """Embed a downscaled JPEG so the self-contained HTML stays small enough to email."""
    try:
        import cv2
        img = cv2.imread(str(path))
        if img is not None:
            h, w = img.shape[:2]
            if w > max_width:
                img = cv2.resize(img, (max_width, int(h * max_width / w)),
                                 interpolation=cv2.INTER_AREA)
            ok, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 70])
            if ok:
                return "data:image/jpeg;base64," + base64.b64encode(buf.tobytes()).decode()
    except Exception:
        pass
    try:
        return "data:image/jpeg;base64," + base64.b64encode(path.read_bytes()).decode()
    except Exception:
        return ""


def build_html(result: ScanResult, wd: Path, out: Path) -> Path:
    devices = result.devices
    total_streams = sum(len(d.streams) for d in devices)
    failures = sum(1 for d in devices if d.status == "failed")

    badge = {"good": "#16a34a", "usable": "#d97706", "failed": "#dc2626"}
    cards = []
    for d in devices:
        rows = []
        for s in d.streams:
            thumbs = "".join(
                f'<img src="{_img_data_uri(wd / sc)}" class="thumb" onclick="zoom(this)">'
                for sc in s.screenshots
            )
            err = f'<div class="err">{s.error}</div>' if s.error else ""
            rows.append(f"""
            <tr>
              <td>{s.profile}</td>
              <td>{s.resolution or '-'}</td>
              <td>{s.codec or '-'}</td>
              <td>{s.fps_advertised or '-'}</td>
              <td><b>{s.fps_measured or '-'}</b></td>
              <td>{s.quality_score}</td>
              <td>{s.exposure or '-'}</td>
              <td class="rtsp"><code>{s.rtsp_url or '-'}</code></td>
              <td>{thumbs}{err}</td>
            </tr>""")
        color = badge.get(d.status, "#6b7280")
        cards.append(f"""
        <div class="card">
          <div class="card-head">
            <span class="dot" style="background:{color}"></span>
            <b>{d.ip}</b> &nbsp; {d.manufacturer} {d.model}
            <span class="muted">{d.firmware}</span>
            <span class="status" style="background:{color}">{d.status.upper()}</span>
          </div>
          <table>
            <tr><th>Profile</th><th>Resolution</th><th>Codec</th><th>FPS adv</th>
                <th>FPS real</th><th>Quality</th><th>Exposure</th><th>RTSP</th><th>Samples</th></tr>
            {''.join(rows) or '<tr><td colspan=9 class="muted">no streams</td></tr>'}
          </table>
          {f'<div class="err">{d.error}</div>' if d.error else ''}
        </div>""")

    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Camera onboarding — {result.client}</title>
<style>
 body{{font-family:-apple-system,Segoe UI,Roboto,sans-serif;margin:0;background:#f8fafc;color:#0f172a}}
 header{{background:#0f172a;color:#fff;padding:18px 24px}}
 header h1{{margin:0;font-size:20px}} .sub{{opacity:.7;font-size:13px;margin-top:4px}}
 .totals{{display:flex;gap:24px;margin-top:12px;font-size:14px}}
 .totals b{{font-size:20px;display:block}}
 .wrap{{padding:20px 24px}}
 .card{{background:#fff;border:1px solid #e2e8f0;border-radius:10px;margin-bottom:16px;overflow:hidden}}
 .card-head{{padding:12px 16px;border-bottom:1px solid #f1f5f9;font-size:15px}}
 .dot{{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:8px}}
 .status{{float:right;color:#fff;font-size:11px;padding:2px 8px;border-radius:10px}}
 .muted{{color:#94a3b8}}
 table{{width:100%;border-collapse:collapse;font-size:13px}}
 th,td{{text-align:left;padding:8px 12px;border-bottom:1px solid #f1f5f9;vertical-align:top}}
 th{{background:#f8fafc;color:#475569;font-weight:600}}
 .rtsp code{{font-size:11px;color:#475569;word-break:break-all}}
 .thumb{{height:64px;border-radius:4px;margin:2px;cursor:zoom-in;border:1px solid #e2e8f0}}
 .err{{color:#dc2626;font-size:12px;margin-top:4px}}
 #lb{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.85);align-items:center;justify-content:center;cursor:zoom-out}}
 #lb img{{max-width:92%;max-height:92%}}
</style></head><body>
<header>
  <h1>Camera onboarding — {result.client or 'Unnamed client'}</h1>
  <div class="sub">Scanned {result.scanned_at}</div>
  <div class="totals">
    <span><b>{len(devices)}</b> devices</span>
    <span><b>{total_streams}</b> streams</span>
    <span><b>{failures}</b> failed</span>
  </div>
</header>
<div class="wrap">{''.join(cards)}</div>
<div id="lb" onclick="this.style.display='none'"><img id="lbimg"></div>
<script>
 function zoom(el){{document.getElementById('lbimg').src=el.src;document.getElementById('lb').style.display='flex';}}
</script>
</body></html>"""
    out.write_text(html)
    return out


# --------------------------------------------------------------------------- Email relay
def email_report(
    result: ScanResult,
    html_path: Path,
    xlsx_path: Path,
    recipients: List[str],
    from_name: str = "",
) -> tuple[bool, str]:
    import requests
    from .config import RELAY_URL, RELAY_TOKEN

    html = html_path.read_text()
    xlsx_b64 = base64.b64encode(xlsx_path.read_bytes()).decode() if xlsx_path.exists() else ""
    payload = {
        "to": recipients,
        "subject": f"Camera onboarding — {result.client or 'client'} — {result.scanned_at[:10]}",
        "from_name": from_name,
        "html": html,
        "xlsx_base64": xlsx_b64,
        "xlsx_name": xlsx_path.name if xlsx_path.exists() else "",
    }
    try:
        r = requests.post(
            RELAY_URL,
            json=payload,
            headers={"Authorization": f"Bearer {RELAY_TOKEN}"},
            timeout=30,
        )
        if r.status_code // 100 == 2:
            return True, "sent"
        return False, f"relay {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, f"relay unreachable: {e}"
